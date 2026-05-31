# encoding:utf-8

"""Local asset catalog for Discord Grok real-mode direct prompts."""

from __future__ import annotations

import json
import os
import random
import tempfile
import threading
import time
from collections.abc import Iterable as IterableABC
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence

from common.log import logger
from common.utils import expand_path
from config import conf


ASSET_FIELDS = (
    "camera_angle",
    "scene",
    "time",
    "light_source",
    "color_tone",
    "nationality",
    "pose",
    "action",
    "clothing",
    "lower_state",
    "tattoo",
    "expression",
)

TATTOO_FIELD = "tattoo"
TATTOO_RANDOM_MANY_KEY = "\u968f\u673a\u591a\u7eb9\u8eab"
TATTOO_RANDOM_FEW_KEY = "\u968f\u673a\u5c11\u7eb9\u8eab"
TATTOO_RANDOM_ALIASES = {
    TATTOO_RANDOM_MANY_KEY: "tattoo_many",
    "random_many": "tattoo_many",
    TATTOO_RANDOM_FEW_KEY: "tattoo_few",
    "random_few": "tattoo_few",
    "random": "tattoo_few",
}
RANDOM_TATTOO_MANY_FIELD = "tattoo_many"
RANDOM_TATTOO_FEW_FIELD = "tattoo_few"
RANDOM_ASSET_FIELDS = tuple(field for field in ASSET_FIELDS if field != TATTOO_FIELD) + (
    RANDOM_TATTOO_MANY_FIELD,
    RANDOM_TATTOO_FEW_FIELD,
)
EXTRA_PROMPT_FIELDS = tuple(f"prompt_{index}" for index in range(2, 7))
DEFAULT_ASSETS_DIR = "data/grok-real-mode-assets"
DEFAULT_ASSETS_XLSX = "grok_real_mode_assets.xlsx"
DEFAULT_ASSETS_CACHE = "grok_real_mode_assets.cache.json"
DEFAULT_RANDOM_ASSETS_XLSX = "grok_real_mode_random_assets.xlsx"
DEFAULT_RANDOM_ASSETS_CACHE = "grok_real_mode_random_assets.cache.json"
DEFAULT_RANDOM_STATE = "grok_real_mode_random_assets.state.json"
CACHE_VERSION = 1
CUSTOM_PREFIXES = ("custom:", "custom\uff1a")
_RANDOM_STATE_LOCK = threading.Lock()

TEXT_TO_IMAGE_TEMPLATE = (
    "Raw hidden iPhone 6s photo from {camera_angle} looking up in {scene} at {time}, "
    "faint {light_source} glow, super grainy high-ISO noise, slight motion blur, "
    "low exposure shadows, {color_tone} cast, shaky amateur hidden camera feel, "
    "One imaginary 20-year-old {nationality} woman (natural attractive features realistic skin pores) "
    "{pose}, {action} wearing {clothing}, {lower_state}, {tattoo_clause}{expression}, {extra_references}"
    "background with {scene}, photorealistic but iPhone 6s low light: heavy digital grain, "
    "soft focus, {color_tone} cast, raw unfiltered iPhone photo style"
)
IMAGE_TO_IMAGE_TEMPLATE = (
    "Raw hidden iPhone 6s photo from {camera_angle} looking up in {scene} at {time}, "
    "faint {light_source} glow, super grainy high-ISO noise, slight motion blur, "
    "low exposure shadows, {color_tone} cast, shaky amateur hidden camera feel, "
    "woman (natural attractive features realistic skin pores) {pose}, {action} wearing {clothing}, "
    "{lower_state}, {tattoo_clause}{expression}, {extra_references}background with {scene}, "
    "photorealistic but iPhone 6s low light: heavy digital grain, soft focus, "
    "{color_tone} cast, raw unfiltered iPhone photo style"
)

DEFAULT_ASSETS: Dict[str, List[Dict[str, str]]] = {
    "camera_angle": [
        {"key": "floor_low", "prompt": "a floor-level hidden low angle"},
        {"key": "table_low", "prompt": "a low table-edge hidden camera angle"},
        {"key": "bag_low", "prompt": "a low hidden bag-camera angle"},
    ],
    "scene": [
        {"key": "bedroom_dim", "prompt": "a dimly lit bedroom"},
        {"key": "hotel_room", "prompt": "a dim hotel room"},
        {"key": "apartment_night", "prompt": "a small apartment room"},
    ],
    "time": [
        {"key": "late_night", "prompt": "late night"},
        {"key": "after_midnight", "prompt": "after midnight"},
        {"key": "blue_hour", "prompt": "blue hour before dawn"},
    ],
    "light_source": [
        {"key": "phone_screen", "prompt": "phone screen"},
        {"key": "laptop_screen", "prompt": "laptop screen"},
        {"key": "warm_lamp", "prompt": "weak warm lamp"},
    ],
    "color_tone": [
        {"key": "warm_amber", "prompt": "warm amber"},
        {"key": "cool_blue", "prompt": "cool blue"},
        {"key": "greenish", "prompt": "greenish fluorescent"},
    ],
    "nationality": [
        {"key": "korean", "prompt": "Korean"},
        {"key": "japanese", "prompt": "Japanese"},
        {"key": "chinese", "prompt": "Chinese"},
        {"key": "american", "prompt": "American"},
    ],
    "pose": [
        {"key": "relaxed_stance", "prompt": "in a relaxed standing pose"},
        {"key": "slight_lean", "prompt": "slightly leaning to one side"},
        {"key": "side_profile", "prompt": "in a soft side-profile pose"},
    ],
    "action": [
        {"key": "standing", "prompt": "standing naturally in a candid pose"},
        {"key": "turning", "prompt": "turning slightly toward the camera"},
        {"key": "walking", "prompt": "walking slowly through the room"},
    ],
    "clothing": [
        {"key": "oversized_hoodie", "prompt": "an oversized hoodie"},
        {"key": "casual_dress", "prompt": "a simple casual dress"},
        {"key": "soft_sweater", "prompt": "a soft loose sweater"},
    ],
    "lower_state": [
        {"key": "casual_shorts", "prompt": "with casual shorts clearly visible"},
        {"key": "long_skirt", "prompt": "with a long skirt clearly visible"},
        {"key": "loose_pants", "prompt": "with loose pants clearly visible"},
    ],
    "tattoo": [
        {"key": "small_wrist", "prompt": "a small subtle wrist tattoo"},
        {"key": "shoulder_flower", "prompt": "a delicate floral shoulder tattoo"},
        {"key": "ankle_line", "prompt": "a minimal fine-line ankle tattoo"},
    ],
    "expression": [
        {"key": "neutral", "prompt": "neutral candid expression"},
        {"key": "slight_smile", "prompt": "a faint natural smile"},
        {"key": "thoughtful", "prompt": "a quiet thoughtful expression"},
    ],
}

DEFAULT_RANDOM_ASSETS: Dict[str, List[Dict[str, str]]] = {
    **{field: [dict(item) for item in DEFAULT_ASSETS[field]] for field in ASSET_FIELDS if field != TATTOO_FIELD},
    RANDOM_TATTOO_MANY_FIELD: [
        {"key": "many_sleeve", "prompt": "multiple visible tattoos across her arms and shoulders"},
        {"key": "many_body", "prompt": "several visible artistic tattoos on her arms, shoulder, and upper chest"},
    ],
    RANDOM_TATTOO_FEW_FIELD: [
        {"key": "small_wrist", "prompt": "a small subtle wrist tattoo"},
        {"key": "ankle_line", "prompt": "a minimal fine-line ankle tattoo"},
    ],
}


class GrokRealModePromptError(ValueError):
    """Raised when a real-mode material selection cannot be resolved."""


@dataclass(frozen=True)
class AssetPaths:
    workbook: Path
    cache: Path


@dataclass(frozen=True)
class RandomAssetPaths:
    workbook: Path
    cache: Path
    state: Path


@dataclass(frozen=True)
class SyncResult:
    status: str
    workbook: Path
    cache: Path
    categories: int


def configured_asset_paths() -> AssetPaths:
    assets_dir = str(conf().get("grok_real_mode_assets_dir") or DEFAULT_ASSETS_DIR).strip()
    workbook_value = str(conf().get("grok_real_mode_assets_xlsx") or "").strip()
    cache_value = str(conf().get("grok_real_mode_assets_cache") or "").strip()
    base_dir = Path(expand_path(assets_dir))
    workbook = Path(expand_path(workbook_value)) if workbook_value else base_dir / DEFAULT_ASSETS_XLSX
    cache = Path(expand_path(cache_value)) if cache_value else base_dir / DEFAULT_ASSETS_CACHE
    return AssetPaths(workbook=workbook.resolve(), cache=cache.resolve())


def configured_random_asset_paths() -> RandomAssetPaths:
    assets_dir = str(conf().get("grok_real_mode_assets_dir") or DEFAULT_ASSETS_DIR).strip()
    workbook_value = str(conf().get("grok_real_mode_random_assets_xlsx") or "").strip()
    cache_value = str(conf().get("grok_real_mode_random_assets_cache") or "").strip()
    state_value = str(conf().get("grok_real_mode_random_state") or "").strip()
    base_dir = Path(expand_path(assets_dir))
    workbook = Path(expand_path(workbook_value)) if workbook_value else base_dir / DEFAULT_RANDOM_ASSETS_XLSX
    cache = Path(expand_path(cache_value)) if cache_value else base_dir / DEFAULT_RANDOM_ASSETS_CACHE
    state = Path(expand_path(state_value)) if state_value else base_dir / DEFAULT_RANDOM_STATE
    return RandomAssetPaths(workbook=workbook.resolve(), cache=cache.resolve(), state=state.resolve())


def sync_workbook_to_cache(
    *,
    workbook_path: Optional[Any] = None,
    cache_path: Optional[Any] = None,
    force: bool = False,
) -> SyncResult:
    paths = _paths_from_args(workbook_path, cache_path)
    paths.workbook.parent.mkdir(parents=True, exist_ok=True)
    paths.cache.parent.mkdir(parents=True, exist_ok=True)
    if not paths.workbook.exists():
        _create_default_workbook(paths.workbook)
    else:
        _ensure_workbook_sheets(paths.workbook, fields=ASSET_FIELDS, defaults=DEFAULT_ASSETS)
    workbook_stat = paths.workbook.stat()
    if not force and paths.cache.exists() and _cache_matches_workbook(paths.cache, workbook_stat):
        return SyncResult("unchanged", paths.workbook, paths.cache, len(ASSET_FIELDS))

    catalog = _read_workbook(paths.workbook)
    payload = {
        "version": CACHE_VERSION,
        "created_at": time.time(),
        "workbook": {
            "path": str(paths.workbook),
            "mtime_ns": workbook_stat.st_mtime_ns,
            "size": workbook_stat.st_size,
        },
        "categories": catalog,
    }
    _write_json_atomic(paths.cache, payload)
    return SyncResult("updated", paths.workbook, paths.cache, len(catalog))


def sync_random_workbook_to_cache(
    *,
    workbook_path: Optional[Any] = None,
    cache_path: Optional[Any] = None,
    force: bool = False,
) -> SyncResult:
    paths = _random_paths_from_args(workbook_path, cache_path, None)
    paths.workbook.parent.mkdir(parents=True, exist_ok=True)
    paths.cache.parent.mkdir(parents=True, exist_ok=True)
    if not paths.workbook.exists():
        _create_default_workbook(paths.workbook, fields=RANDOM_ASSET_FIELDS, defaults=DEFAULT_RANDOM_ASSETS)
    else:
        _ensure_workbook_sheets(paths.workbook, fields=RANDOM_ASSET_FIELDS, defaults=DEFAULT_RANDOM_ASSETS)
    workbook_stat = paths.workbook.stat()
    if not force and paths.cache.exists() and _cache_matches_workbook(paths.cache, workbook_stat):
        return SyncResult("unchanged", paths.workbook, paths.cache, len(RANDOM_ASSET_FIELDS))

    catalog = _read_workbook(paths.workbook, fields=RANDOM_ASSET_FIELDS, defaults=DEFAULT_RANDOM_ASSETS)
    payload = {
        "version": CACHE_VERSION,
        "created_at": time.time(),
        "workbook": {
            "path": str(paths.workbook),
            "mtime_ns": workbook_stat.st_mtime_ns,
            "size": workbook_stat.st_size,
        },
        "categories": catalog,
    }
    _write_json_atomic(paths.cache, payload)
    return SyncResult("updated", paths.workbook, paths.cache, len(catalog))


def load_catalog(*, cache_path: Optional[Any] = None) -> Dict[str, List[Dict[str, str]]]:
    paths = _paths_from_args(None, cache_path)
    try:
        with paths.cache.open("r", encoding="utf-8") as fh:
            payload = json.load(fh)
        categories = payload.get("categories") if isinstance(payload, dict) else None
        return _normalize_catalog(categories)
    except Exception as exc:
        logger.debug("[GrokRealMode] using default catalog; cache load failed: %s", exc)
        return _default_catalog()


def load_random_catalog(*, cache_path: Optional[Any] = None) -> Dict[str, List[Dict[str, str]]]:
    paths = _random_paths_from_args(None, cache_path, None)
    try:
        with paths.cache.open("r", encoding="utf-8") as fh:
            payload = json.load(fh)
        categories = payload.get("categories") if isinstance(payload, dict) else None
        return _normalize_random_catalog(categories)
    except Exception as exc:
        logger.debug("[GrokRealMode] using default random catalog; cache load failed: %s", exc)
        return _default_random_catalog()


def material_choices(field: str, current: str = "", *, limit: int = 25) -> List[str]:
    field = _normalize_field_name(field)
    needle = str(current or "").strip().lower()
    if needle.startswith(CUSTOM_PREFIXES):
        return []
    catalog = load_catalog()
    choices: List[str] = []
    if field == TATTOO_FIELD:
        special_choices = (
            (TATTOO_RANDOM_MANY_KEY, "random_many"),
            (TATTOO_RANDOM_FEW_KEY, "random_few"),
        )
        for special_key, alias in special_choices:
            if not needle or special_key.lower().startswith(needle) or alias.startswith(needle):
                choices.append(special_key)
                if len(choices) >= limit:
                    return choices
    for item in catalog.get(field, []):
        key = str(item.get("key") or "").strip()
        if not key or len(key) > 100:
            continue
        if needle and needle not in key.lower():
            continue
        if field == TATTOO_FIELD and key in TATTOO_RANDOM_ALIASES:
            continue
        choices.append(key)
        if len(choices) >= limit:
            break
    return choices


def compose_real_mode_prompt(
    *,
    media_type: str,
    image_count: int,
    selections: Optional[Mapping[str, Any]] = None,
    extra_prompts: Optional[Mapping[int, Any]] = None,
    catalog: Optional[Mapping[str, Sequence[Mapping[str, str]]]] = None,
    random_catalog: Optional[Mapping[str, Sequence[Mapping[str, str]]]] = None,
    random_state_path: Optional[Any] = None,
    rng: Optional[random.Random] = None,
) -> str:
    selected = selections or {}
    catalog_data = _normalize_catalog(catalog) if catalog is not None else load_catalog()
    random_catalog_data = (
        _normalize_random_catalog(random_catalog)
        if random_catalog is not None
        else (_normalize_random_catalog(catalog) if catalog is not None else load_random_catalog())
    )
    values = {
        field: _resolve_material(
            field,
            selected.get(field),
            catalog_data,
            random_catalog=random_catalog_data,
            random_state_path=random_state_path,
            rng=rng,
        )
        for field in ASSET_FIELDS
    }
    values["tattoo_clause"] = f"{values[TATTOO_FIELD]}, " if values.get(TATTOO_FIELD) else ""
    values["extra_references"] = _extra_reference_clause(
        media_type=media_type,
        image_count=max(0, int(image_count or 0)),
        extra_prompts=extra_prompts or {},
    )
    template = TEXT_TO_IMAGE_TEMPLATE if max(0, int(image_count or 0)) == 0 else IMAGE_TO_IMAGE_TEMPLATE
    return template.format(**values)


def _paths_from_args(workbook_path: Optional[Any], cache_path: Optional[Any]) -> AssetPaths:
    configured = configured_asset_paths()
    workbook = Path(expand_path(str(workbook_path))).resolve() if workbook_path else configured.workbook
    cache = Path(expand_path(str(cache_path))).resolve() if cache_path else configured.cache
    return AssetPaths(workbook=workbook, cache=cache)


def _random_paths_from_args(
    workbook_path: Optional[Any],
    cache_path: Optional[Any],
    state_path: Optional[Any],
) -> RandomAssetPaths:
    configured = configured_random_asset_paths()
    workbook = Path(expand_path(str(workbook_path))).resolve() if workbook_path else configured.workbook
    cache = Path(expand_path(str(cache_path))).resolve() if cache_path else configured.cache
    state = Path(expand_path(str(state_path))).resolve() if state_path else configured.state
    return RandomAssetPaths(workbook=workbook, cache=cache, state=state)


def _create_default_workbook(
    path: Path,
    *,
    fields: Sequence[str] = ASSET_FIELDS,
    defaults: Mapping[str, Sequence[Mapping[str, str]]] = DEFAULT_ASSETS,
) -> None:
    try:
        from openpyxl import Workbook
    except Exception as exc:  # pragma: no cover - depends on deployment deps
        raise GrokRealModePromptError("openpyxl is required to create the Grok real-mode assets workbook.") from exc

    wb = Workbook()
    default_sheet = wb.active
    default_sheet.title = fields[0]
    for field_index, field in enumerate(fields):
        ws = default_sheet if field_index == 0 else wb.create_sheet(field)
        for row_index, item in enumerate(defaults[field], start=1):
            ws.cell(row=row_index, column=1, value=item["key"])
            ws.cell(row=row_index, column=2, value=item["prompt"])
    wb.save(path)


def _ensure_workbook_sheets(
    path: Path,
    *,
    fields: Sequence[str],
    defaults: Mapping[str, Sequence[Mapping[str, str]]],
) -> None:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - depends on deployment deps
        raise GrokRealModePromptError("openpyxl is required to update Grok real-mode assets.") from exc

    wb = load_workbook(path)
    try:
        changed = False
        for field_index, field in enumerate(fields):
            if field in wb.sheetnames:
                continue
            ws = wb.create_sheet(field, index=field_index)
            for row_index, item in enumerate(defaults[field], start=1):
                ws.cell(row=row_index, column=1, value=item["key"])
                ws.cell(row=row_index, column=2, value=item["prompt"])
            changed = True
        if changed:
            wb.save(path)
    finally:
        wb.close()


def _read_workbook(
    path: Path,
    *,
    fields: Sequence[str] = ASSET_FIELDS,
    defaults: Mapping[str, Sequence[Mapping[str, str]]] = DEFAULT_ASSETS,
) -> Dict[str, List[Dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - depends on deployment deps
        raise GrokRealModePromptError("openpyxl is required to read Grok real-mode assets.") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        catalog: Dict[str, List[Dict[str, str]]] = {}
        for field in fields:
            if field not in wb.sheetnames:
                catalog[field] = [dict(item) for item in defaults[field]]
                continue
            ws = wb[field]
            seen: set[str] = set()
            items: List[Dict[str, str]] = []
            for row in ws.iter_rows(min_row=1, values_only=True):
                key = str(row[0] or "").strip() if len(row) >= 1 else ""
                prompt = str(row[1] or "").strip() if len(row) >= 2 else ""
                if not key and not prompt:
                    continue
                if not key or not prompt:
                    raise GrokRealModePromptError(f"Sheet {field} contains an incomplete A/B row.")
                if key in seen:
                    raise GrokRealModePromptError(f"Sheet {field} contains duplicate material key: {key}")
                seen.add(key)
                items.append({"key": key, "prompt": prompt})
            catalog[field] = items or [dict(item) for item in defaults[field]]
        return catalog
    finally:
        wb.close()


def _cache_matches_workbook(cache_path: Path, workbook_stat: os.stat_result) -> bool:
    try:
        with cache_path.open("r", encoding="utf-8") as fh:
            payload = json.load(fh)
        workbook = payload.get("workbook") if isinstance(payload, dict) else None
        return (
            isinstance(workbook, dict)
            and int(workbook.get("mtime_ns") or -1) == int(workbook_stat.st_mtime_ns)
            and int(workbook.get("size") or -1) == int(workbook_stat.st_size)
        )
    except Exception:
        return False


def _write_json_atomic(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=str(path.parent))
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, ensure_ascii=False, indent=2)
            fh.write("\n")
        os.replace(tmp_name, path)
    finally:
        if os.path.exists(tmp_name):
            os.remove(tmp_name)


def _normalize_catalog(value: Any) -> Dict[str, List[Dict[str, str]]]:
    catalog = _default_catalog()
    if not isinstance(value, Mapping):
        return catalog
    for field in ASSET_FIELDS:
        items = value.get(field)
        normalized: List[Dict[str, str]] = []
        if isinstance(items, IterableABC) and not isinstance(items, (str, bytes, dict)):
            for item in items:
                if not isinstance(item, Mapping):
                    continue
                key = str(item.get("key") or "").strip()
                prompt = str(item.get("prompt") or "").strip()
                if key and prompt:
                    normalized.append({"key": key, "prompt": prompt})
        if normalized:
            catalog[field] = normalized
    return catalog


def _normalize_random_catalog(value: Any) -> Dict[str, List[Dict[str, str]]]:
    catalog = _default_random_catalog()
    if not isinstance(value, Mapping):
        return catalog
    for field in RANDOM_ASSET_FIELDS:
        items = value.get(field)
        normalized: List[Dict[str, str]] = []
        if isinstance(items, IterableABC) and not isinstance(items, (str, bytes, dict)):
            for index, item in enumerate(items, start=1):
                if not isinstance(item, Mapping):
                    continue
                key = str(item.get("key") or "").strip() or f"{field}_{index}"
                prompt = str(item.get("prompt") or "").strip()
                if key and prompt:
                    normalized.append({"key": key, "prompt": prompt})
        if normalized:
            catalog[field] = normalized
    return catalog


def _default_catalog() -> Dict[str, List[Dict[str, str]]]:
    return {field: [dict(item) for item in DEFAULT_ASSETS[field]] for field in ASSET_FIELDS}


def _default_random_catalog() -> Dict[str, List[Dict[str, str]]]:
    return {field: [dict(item) for item in DEFAULT_RANDOM_ASSETS[field]] for field in RANDOM_ASSET_FIELDS}


def _resolve_material(
    field: str,
    selection: Any,
    catalog: Mapping[str, Sequence[Mapping[str, str]]],
    *,
    random_catalog: Mapping[str, Sequence[Mapping[str, str]]],
    random_state_path: Optional[Any] = None,
    rng: Optional[random.Random] = None,
) -> str:
    field = _normalize_field_name(field)
    raw = str(selection or "").strip()
    lowered = raw.lower()
    for prefix in CUSTOM_PREFIXES:
        if lowered.startswith(prefix):
            custom = raw[len(prefix) :].strip()
            if custom:
                return custom
            raise GrokRealModePromptError(f"{field} custom value is empty.")

    items = list(catalog.get(field) or DEFAULT_ASSETS[field])
    if not raw:
        if field == TATTOO_FIELD:
            return ""
        return _choose_random_material(
            field,
            random_catalog=random_catalog,
            fallback_items=items,
            state_path=random_state_path,
            rng=rng,
        )
    tattoo_pool = TATTOO_RANDOM_ALIASES.get(raw)
    if field == TATTOO_FIELD and tattoo_pool:
        return _choose_random_material(
            tattoo_pool,
            random_catalog=random_catalog,
            fallback_items=items,
            state_path=random_state_path,
            rng=rng,
        )
    for item in items:
        if str(item.get("key") or "").strip() == raw:
            return str(item.get("prompt") or "").strip()
    raise GrokRealModePromptError(f"Unknown {field} material key: {raw}. Use autocomplete or custom:<prompt>.")


def _choose_random_material(
    field: str,
    *,
    random_catalog: Mapping[str, Sequence[Mapping[str, str]]],
    fallback_items: Sequence[Mapping[str, str]],
    state_path: Optional[Any],
    rng: Optional[random.Random],
) -> str:
    items = list(random_catalog.get(field) or fallback_items)
    if not items:
        raise GrokRealModePromptError(f"No random material available for {field}.")
    if state_path is None:
        state_path = configured_random_asset_paths().state
    state_file = Path(expand_path(str(state_path))).resolve()
    with _RANDOM_STATE_LOCK:
        state = _read_random_state(state_file)
        selected_key = _next_random_key(field, items, state, rng=rng)
        _write_json_atomic(state_file, state)
    for item in items:
        if str(item.get("key") or "").strip() == selected_key:
            return str(item.get("prompt") or "").strip()
    return str(items[0].get("prompt") or "").strip()


def _read_random_state(path: Path) -> Dict[str, Any]:
    try:
        with path.open("r", encoding="utf-8") as fh:
            payload = json.load(fh)
        if isinstance(payload, dict):
            payload.setdefault("version", CACHE_VERSION)
            payload.setdefault("queues", {})
            payload.setdefault("last", {})
            return payload
    except Exception:
        pass
    return {"version": CACHE_VERSION, "queues": {}, "last": {}}


def _next_random_key(
    field: str,
    items: Sequence[Mapping[str, str]],
    state: Dict[str, Any],
    *,
    rng: Optional[random.Random],
) -> str:
    key_to_item = {str(item.get("key") or "").strip(): item for item in items if str(item.get("key") or "").strip()}
    keys = list(key_to_item)
    if not keys:
        raise GrokRealModePromptError(f"No random material available for {field}.")
    queues = state.setdefault("queues", {})
    last_values = state.setdefault("last", {})
    queue = [key for key in queues.get(field, []) if key in key_to_item]
    if not queue:
        queue = list(keys)
        chooser = rng if rng is not None else random
        chooser.shuffle(queue)
        last_key = str(last_values.get(field) or "")
        if len(queue) > 1 and queue[0] == last_key:
            queue.append(queue.pop(0))
    selected_key = queue.pop(0)
    queues[field] = queue
    last_values[field] = selected_key
    return selected_key


def _normalize_field_name(field: str) -> str:
    normalized = str(field or "").strip()
    if normalized not in ASSET_FIELDS:
        raise GrokRealModePromptError(f"Unknown Grok real-mode material field: {normalized}")
    return normalized


def _extra_reference_clause(*, media_type: str, image_count: int, extra_prompts: Mapping[int, Any]) -> str:
    if image_count <= 1:
        return ""
    fragments: List[str] = []
    for index in range(2, image_count + 1):
        text = str(extra_prompts.get(index) or "").strip()
        if not text:
            raise GrokRealModePromptError(f"image{index} requires prompt_{index} in real mode.")
        if str(media_type or "").strip().lower() == "video":
            fragments.append(f"<IMAGE_{index}> is {text}")
        else:
            fragments.append(f"Reference image {index} is {text}")
    return ", ".join(fragments) + ", "
