import json
import random
import tempfile
import unittest
from contextlib import contextmanager
from pathlib import Path

from config import conf
from common.grok_real_mode_prompt_assets import (
    ASSET_FIELDS,
    COLOR_TONE_RANDOM_KEY,
    EXPRESSION_RANDOM_KEY,
    GrokRealModePromptError,
    RANDOM_ASSET_FIELDS,
    TATTOO_RANDOM_FEW_KEY,
    TATTOO_RANDOM_MANY_KEY,
    compose_real_mode_prompt,
    load_catalog,
    load_random_catalog,
    material_choices,
    sync_random_workbook_to_cache,
    sync_workbook_to_cache,
)


@contextmanager
def patch_config(values):
    old_values = {key: conf().get(key) for key in values}
    conf().update(values)
    try:
        yield
    finally:
        for key, old_value in old_values.items():
            if old_value is None:
                conf().pop(key, None)
            else:
                conf()[key] = old_value


def _single_item_catalog():
    return {field: [{"key": "one", "prompt": f"{field} prompt"}] for field in ASSET_FIELDS}


def _single_item_random_catalog():
    return {field: [{"key": "one", "prompt": f"{field} prompt"}] for field in RANDOM_ASSET_FIELDS}


class GrokRealModePromptAssetsTest(unittest.TestCase):
    def test_sync_creates_default_workbook_and_cache(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = root / "assets.xlsx"
            cache = root / "assets.json"

            first = sync_workbook_to_cache(workbook_path=workbook, cache_path=cache)
            second = sync_workbook_to_cache(workbook_path=workbook, cache_path=cache)

            self.assertEqual(first.status, "updated")
            self.assertEqual(second.status, "unchanged")
            self.assertTrue(workbook.exists())
            self.assertTrue(cache.exists())
            payload = json.loads(cache.read_text(encoding="utf-8"))
            self.assertEqual(set(payload["categories"]), set(ASSET_FIELDS))

    def test_sync_adds_missing_pose_sheet_without_rebuilding_existing_workbook(self):
        from openpyxl import Workbook, load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = root / "assets.xlsx"
            cache = root / "assets.json"
            wb = Workbook()
            ws = wb.active
            ws.title = "camera_angle"
            ws.cell(row=1, column=1, value="kept")
            ws.cell(row=1, column=2, value="kept camera prompt")
            wb.save(workbook)

            sync_workbook_to_cache(workbook_path=workbook, cache_path=cache)

            reloaded = load_workbook(workbook, read_only=True)
            try:
                self.assertIn("pose", reloaded.sheetnames)
                self.assertLess(reloaded.sheetnames.index("pose"), reloaded.sheetnames.index("action"))
                self.assertEqual(reloaded["camera_angle"].cell(row=1, column=1).value, "kept")
                self.assertEqual(reloaded["camera_angle"].cell(row=1, column=2).value, "kept camera prompt")
            finally:
                reloaded.close()

    def test_sync_creates_default_random_workbook_and_cache(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = root / "random-assets.xlsx"
            cache = root / "random-assets.json"

            first = sync_random_workbook_to_cache(workbook_path=workbook, cache_path=cache)
            second = sync_random_workbook_to_cache(workbook_path=workbook, cache_path=cache)

            self.assertEqual(first.status, "updated")
            self.assertEqual(second.status, "unchanged")
            self.assertTrue(workbook.exists())
            self.assertTrue(cache.exists())
            payload = json.loads(cache.read_text(encoding="utf-8"))
            self.assertEqual(set(payload["categories"]), set(RANDOM_ASSET_FIELDS))
            self.assertIn("tattoo_many", payload["categories"])
            self.assertIn("tattoo_few", payload["categories"])

    def test_load_catalog_uses_default_when_cache_is_missing(self):
        with tempfile.TemporaryDirectory() as tmp:
            catalog = load_catalog(cache_path=Path(tmp) / "missing.json")

        self.assertIn("camera_angle", catalog)
        self.assertTrue(catalog["camera_angle"])

    def test_load_random_catalog_uses_default_when_cache_is_missing(self):
        with tempfile.TemporaryDirectory() as tmp:
            catalog = load_random_catalog(cache_path=Path(tmp) / "missing.json")

        self.assertIn("camera_angle", catalog)
        self.assertIn("tattoo_many", catalog)
        self.assertTrue(catalog["camera_angle"])

    def test_compose_text_to_image_prompt_uses_custom_materials(self):
        prompt = compose_real_mode_prompt(
            media_type="image",
            image_count=0,
            selections={
                "camera_angle": "custom:under a chair",
                "scene": "custom:attic room",
                "time": "custom:midnight",
                "light_source": "custom:phone",
                "color_tone": "custom:violet",
                "nationality": "custom:French",
                "pose": "custom:in a seated pose",
                "action": "custom:looking back",
                "clothing": "custom:a wool coat",
                "lower_state": "custom:with black trousers",
                "tattoo": "custom:a tiny star tattoo",
                "expression": "custom:soft expression",
            },
            catalog=_single_item_catalog(),
        )

        self.assertIn("One imaginary 20-year-old French woman", prompt)
        self.assertLess(prompt.index("in a seated pose"), prompt.index("looking back"))
        self.assertLess(prompt.index("a tiny star tattoo"), prompt.index("soft expression"))
        self.assertIn("background with attic room", prompt)

    def test_tattoo_is_omitted_by_default(self):
        prompt = compose_real_mode_prompt(
            media_type="image",
            image_count=0,
            selections={field: "one" for field in ASSET_FIELDS if field != "tattoo"},
            catalog=_single_item_catalog(),
        )

        self.assertNotIn("tattoo prompt", prompt)
        self.assertIn("expression prompt", prompt)

    def test_expression_is_omitted_by_default(self):
        prompt = compose_real_mode_prompt(
            media_type="image",
            image_count=0,
            selections={field: "one" for field in ASSET_FIELDS if field != "expression"},
            catalog=_single_item_catalog(),
        )

        self.assertIn("lower_state prompt", prompt)
        self.assertNotIn("expression prompt", prompt)
        self.assertNotIn(", ,", prompt)

    def test_color_tone_is_omitted_by_default(self):
        prompt = compose_real_mode_prompt(
            media_type="image",
            image_count=0,
            selections={field: "one" for field in ASSET_FIELDS if field != "color_tone"},
            catalog=_single_item_catalog(),
        )

        self.assertNotIn("color_tone prompt", prompt)
        self.assertIn("low exposure shadows, shaky amateur", prompt)
        self.assertIn("soft focus, raw unfiltered", prompt)

    def test_tattoo_random_option_uses_catalog_material(self):
        prompt = compose_real_mode_prompt(
            media_type="image",
            image_count=0,
            selections={**{field: "one" for field in ASSET_FIELDS if field != "tattoo"}, "tattoo": "random"},
            catalog=_single_item_catalog(),
            random_catalog=_single_item_random_catalog(),
            rng=random.Random(0),
        )

        self.assertIn("tattoo_few prompt", prompt)
        self.assertLess(prompt.index("tattoo_few prompt"), prompt.index("expression prompt"))

    def test_tattoo_many_and_few_random_options_use_separate_random_sheets(self):
        selections = {field: "one" for field in ASSET_FIELDS if field != "tattoo"}
        random_catalog = _single_item_random_catalog()

        many_prompt = compose_real_mode_prompt(
            media_type="image",
            image_count=0,
            selections={**selections, "tattoo": TATTOO_RANDOM_MANY_KEY},
            catalog=_single_item_catalog(),
            random_catalog=random_catalog,
            rng=random.Random(0),
        )
        few_prompt = compose_real_mode_prompt(
            media_type="image",
            image_count=0,
            selections={**selections, "tattoo": TATTOO_RANDOM_FEW_KEY},
            catalog=_single_item_catalog(),
            random_catalog=random_catalog,
            rng=random.Random(0),
        )

        self.assertIn("tattoo_many prompt", many_prompt)
        self.assertNotIn("tattoo_few prompt", many_prompt)
        self.assertIn("tattoo_few prompt", few_prompt)
        self.assertNotIn("tattoo_many prompt", few_prompt)

    def test_expression_random_option_uses_random_catalog_material(self):
        prompt = compose_real_mode_prompt(
            media_type="image",
            image_count=0,
            selections={
                **{field: "one" for field in ASSET_FIELDS if field not in {"expression", "tattoo"}},
                "expression": EXPRESSION_RANDOM_KEY,
            },
            catalog=_single_item_catalog(),
            random_catalog=_single_item_random_catalog(),
            rng=random.Random(0),
        )

        self.assertIn("expression prompt", prompt)
        self.assertNotIn("tattoo prompt", prompt)
        self.assertIn("lower_state prompt, expression prompt", prompt)

    def test_color_tone_random_option_uses_random_catalog_material(self):
        prompt = compose_real_mode_prompt(
            media_type="image",
            image_count=0,
            selections={
                **{field: "one" for field in ASSET_FIELDS if field not in {"color_tone", "expression", "tattoo"}},
                "color_tone": COLOR_TONE_RANDOM_KEY,
            },
            catalog=_single_item_catalog(),
            random_catalog=_single_item_random_catalog(),
            rng=random.Random(0),
        )

        self.assertEqual(prompt.count("color_tone prompt cast"), 2)
        self.assertIn("low exposure shadows, color_tone prompt cast", prompt)
        self.assertNotIn("tattoo prompt", prompt)
        self.assertNotIn("expression prompt", prompt)

    def test_empty_selection_uses_large_random_catalog_instead_of_visible_catalog(self):
        random_catalog = {
            "camera_angle": [{"key": "random_angle", "prompt": "large random camera angle"}],
        }
        selections = {field: "one" for field in ASSET_FIELDS if field != "camera_angle"}

        prompt = compose_real_mode_prompt(
            media_type="image",
            image_count=0,
            selections=selections,
            catalog=_single_item_catalog(),
            random_catalog=random_catalog,
            rng=random.Random(0),
        )

        self.assertIn("from large random camera angle looking up", prompt)
        self.assertNotIn("from camera_angle prompt looking up", prompt)

    def test_random_pool_state_rotates_without_repeating_until_exhausted(self):
        random_catalog = {
            "camera_angle": [
                {"key": "angle_a", "prompt": "random angle A"},
                {"key": "angle_b", "prompt": "random angle B"},
            ],
        }
        selections = {field: "one" for field in ASSET_FIELDS if field != "camera_angle"}
        with tempfile.TemporaryDirectory() as tmp:
            state = Path(tmp) / "state.json"
            first = compose_real_mode_prompt(
                media_type="image",
                image_count=0,
                selections=selections,
                catalog=_single_item_catalog(),
                random_catalog=random_catalog,
                random_state_path=state,
                rng=random.Random(7),
            )
            second = compose_real_mode_prompt(
                media_type="image",
                image_count=0,
                selections=selections,
                catalog=_single_item_catalog(),
                random_catalog=random_catalog,
                random_state_path=state,
                rng=random.Random(7),
            )

        self.assertNotEqual("random angle A" in first, "random angle A" in second)
        self.assertTrue(("random angle A" in first) or ("random angle B" in first))
        self.assertTrue(("random angle A" in second) or ("random angle B" in second))

    def test_compose_image_to_image_prompt_inserts_extra_reference_roles(self):
        prompt = compose_real_mode_prompt(
            media_type="image",
            image_count=3,
            selections={field: "one" for field in ASSET_FIELDS},
            extra_prompts={2: "the city background", 3: "the color palette"},
            catalog=_single_item_catalog(),
        )

        self.assertNotIn("One imaginary 20-year-old", prompt)
        self.assertIn("Reference image 2 is the city background", prompt)
        self.assertIn("Reference image 3 is the color palette", prompt)
        self.assertLess(prompt.index("Reference image 2"), prompt.index("background with scene prompt"))

    def test_compose_video_prompt_uses_image_markers_for_extra_references(self):
        prompt = compose_real_mode_prompt(
            media_type="video",
            image_count=2,
            selections={field: "one" for field in ASSET_FIELDS},
            extra_prompts={2: "the motion background"},
            catalog=_single_item_catalog(),
        )

        self.assertIn("<IMAGE_2> is the motion background", prompt)

    def test_compose_requires_extra_prompt_for_each_extra_reference(self):
        with self.assertRaises(GrokRealModePromptError):
            compose_real_mode_prompt(
                media_type="image",
                image_count=2,
                selections={field: "one" for field in ASSET_FIELDS},
                extra_prompts={},
                catalog=_single_item_catalog(),
            )

    def test_compose_rejects_unknown_material_key(self):
        with self.assertRaises(GrokRealModePromptError):
            compose_real_mode_prompt(
                media_type="image",
                image_count=0,
                selections={"camera_angle": "missing"},
                catalog=_single_item_catalog(),
            )

    def test_empty_selection_randomly_uses_catalog_material(self):
        prompt = compose_real_mode_prompt(
            media_type="image",
            image_count=0,
            selections={},
            catalog=_single_item_catalog(),
            random_catalog=_single_item_random_catalog(),
            rng=random.Random(0),
        )

        self.assertIn("camera_angle prompt", prompt)
        self.assertIn("scene prompt", prompt)
        self.assertNotIn("color_tone prompt", prompt)
        self.assertNotIn("tattoo prompt", prompt)
        self.assertNotIn("expression prompt", prompt)

    def test_tattoo_choices_include_random_option(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            cache = root / "assets.json"
            payload = {
                "categories": {
                    field: [{"key": "one", "prompt": f"{field} prompt"}]
                    for field in ASSET_FIELDS
                }
            }
            cache.write_text(json.dumps(payload), encoding="utf-8")

            with patch_config({"grok_real_mode_assets_cache": str(cache)}):
                choices = material_choices("tattoo", "")
                random_prefix_choices = material_choices("tattoo", "\u968f\u673a\u591a")
                english_alias_choices = material_choices("tattoo", "random_f")
                unrelated_choices = material_choices("tattoo", "dom")

        self.assertEqual(choices[:3], [TATTOO_RANDOM_MANY_KEY, TATTOO_RANDOM_FEW_KEY, "one"])
        self.assertEqual(random_prefix_choices, [TATTOO_RANDOM_MANY_KEY])
        self.assertEqual(english_alias_choices, [TATTOO_RANDOM_FEW_KEY])
        self.assertNotIn(TATTOO_RANDOM_MANY_KEY, unrelated_choices)
        self.assertNotIn(TATTOO_RANDOM_FEW_KEY, unrelated_choices)

    def test_expression_choices_include_random_option(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            cache = root / "assets.json"
            payload = {
                "categories": {
                    field: [{"key": "one", "prompt": f"{field} prompt"}]
                    for field in ASSET_FIELDS
                }
            }
            cache.write_text(json.dumps(payload), encoding="utf-8")

            with patch_config({"grok_real_mode_assets_cache": str(cache)}):
                choices = material_choices("expression", "")
                random_alias_choices = material_choices("expression", "random")
                unrelated_choices = material_choices("expression", "zzz")

        self.assertEqual(choices[:2], [EXPRESSION_RANDOM_KEY, "one"])
        self.assertEqual(random_alias_choices, [EXPRESSION_RANDOM_KEY])
        self.assertNotIn(EXPRESSION_RANDOM_KEY, unrelated_choices)

    def test_color_tone_choices_include_random_option(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            cache = root / "assets.json"
            payload = {
                "categories": {
                    field: [{"key": "one", "prompt": f"{field} prompt"}]
                    for field in ASSET_FIELDS
                }
            }
            cache.write_text(json.dumps(payload), encoding="utf-8")

            with patch_config({"grok_real_mode_assets_cache": str(cache)}):
                choices = material_choices("color_tone", "")
                random_alias_choices = material_choices("color_tone", "random")
                unrelated_choices = material_choices("color_tone", "zzz")

        self.assertEqual(choices[:2], [COLOR_TONE_RANDOM_KEY, "one"])
        self.assertEqual(random_alias_choices, [COLOR_TONE_RANDOM_KEY])
        self.assertNotIn(COLOR_TONE_RANDOM_KEY, unrelated_choices)


if __name__ == "__main__":
    unittest.main()
